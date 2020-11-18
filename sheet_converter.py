# -*- coding: UTF-8 -*-
from openpyxl import load_workbook, Workbook
import json
from convertBugzbook import createXMLFile, bugzbookParser
import traceback
import os


def sheet_to_xml_converter(path_to_sheet, path_to_score, repo_name):
    '''
    convert a sheet (format given by HUAWEI) to xml format.
    return the path to store xml
    '''
    workbook = load_workbook(path_to_sheet)
    booksheet = workbook.active
    rows = booksheet.rows
    bug_report_data = {}
    for row in list(rows)[1:]:
        line = [col.value for col in row]
        # line[0]是ID
        id = str(line[0])
        summaryText = line[1]
        linked_file = line[6]
        descriptionText = line[5]
        opendate = line[7]
        if id in bug_report_data.keys():
            bug_report_data[id]["files"].append(linked_file)
        else:
            bug_report_data[id] = {}
            bug_report_data[id]["id"] = id
            bug_report_data[id]["files"] = [linked_file]
            bug_report_data[id]['summary'] = summaryText
            bug_report_data[id]['description'] = descriptionText 
            bug_report_data[id]['opendate'] = opendate
            bug_report_data[id]['fixdate'] = opendate 
            # Missing for fixdate data in the sheet, relace with open date
    allBugData = {}
    allBugData['repositoryName'] = repo_name
    allBugData['bugReportData'] = bug_report_data
    return createXMLFile(allBugData, path_to_score)

def xml_to_sheet_converter(path_to_xml, path_to_save, verbose=False):
    '''
    这个函数主要适用于测试sheet_to_xml_converter()
    '''
    workbook = Workbook()
    booksheet = workbook.active
    booksheet.append(["ISSUENO","DISCRIPT","CREATER","SDEPT3SUBMIT","BVERSION","DETAILDESCRIPTION","FILE_PATH","CREATE_TIME","FEATURE","MODULE"])
    all_bug_data = bugzbookParser(path_to_xml)
    repo_name = all_bug_data["repositoryName"]
    for bug_id in all_bug_data['bugReportData'].keys():
        for linked_file in all_bug_data['bugReportData'][bug_id]['files']:
            summaryText = all_bug_data['bugReportData'][bug_id]['summary']
            descriptionText = all_bug_data['bugReportData'][bug_id]['description']
            creater = 'Creator'
            SDEPT3SUBMIT = 'SDEPT3SUBMIT'
            BVERSION = "BVERSION"
            CREATE_TIME = all_bug_data['bugReportData'][bug_id]['opendate']
            FEATURE = "FEATURE"
            MODULE = "MODULE"
            booksheet.append([bug_id, summaryText, creater, SDEPT3SUBMIT, BVERSION, descriptionText, linked_file, CREATE_TIME, FEATURE, MODULE])
    print("The xlsx file is stored in {}".format(os.path.join(path_to_save, repo_name + '.xlsx')))
    workbook.save(os.path.join(path_to_save, repo_name + '.xlsx'))
    return path_to_save





if __name__=='__main__':
    root_path = '/media/zyang/bugzbook_data/'
    path_to_save = os.path.join(root_path, 'sheet')
    if not os.path.exists(path_to_save):
        os.mkdir(path_to_save)
    for repo_name in os.listdir(os.path.join(root_path, "bugreport_data")):
        if repo_name == '.DS_Store':
            continue
        path_to_xml = os.path.join(root_path, "new_merged_report", repo_name + ".XML")

        xml_to_sheet_converter(path_to_xml, path_to_save, True)

